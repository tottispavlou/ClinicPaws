import os
import PyPDF2
import openpyxl
import sys

bloodwork_path = sys.argv[1]
patient = sys.argv[2]

file_name = os.path.basename(bloodwork_path)
cdir = os.path.dirname(__file__)

if patient:
    
  #  folder_path = fr'C:\Users\User\Documents\Denmark\DTU\semester1\2830\ClinicPaws-stella_new_1\patient_data\{patient}'
    file_path = os.path.join(cdir, "patient_data", patient, file_name.replace('.pdf', '.xlsx'))

    with open(bloodwork_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        totalPages = len(reader.pages)
        pdfData = []
        line = ""
        Tests = ["RBC","HCT","HGB","MCV","MCH","MCHC","RDW","%RETIC","RETIC","RETIC-HGB","WBC","%NEU","%L","%BASO","%EOS","NEU","LYM","MONO","EOS","BASO","PLT","MPV","PDW","PCT","GLU","CREA","BUN","BUN/CREA","PHOS","CA","TP","ALB","GLOB","ALB/GLOB","ALT","ALKP","GGT","TBIL","CHOL","AMYL","LIPA","Na+","K+","Cl-","Ca++","Glu","Lac"]
        flagw = 1
        flags = 0
        cnt = 0
        Res = ""
        RefV = ""
        arrayData = []

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        
        for i in range(len(reader.pages)):
            page = reader.pages[i]
            for char in page.extract_text():
                line += char
                if char == '\n':

                    for j in Tests:
                        if j in line:
                            cnt += 1
                            line.replace(" ", "")
                            for k in line:
                                if ((k >= 'A' and k <= 'Z') or (k >= 'a' and k <= 'z')) and flags == 0:
                                    flagw = 0
                                elif ((k >= '0' and k <= '9') or (k == '.')) and flagw == 0:
                                    Res += k
                                    flags = 1
                                elif ((k >= 'A' and k <= 'Z') or (k >= 'a' and k <= 'z') or k == '%') and flags == 1:
                                    flagw = 1
                                elif ((k >= '0' and k <= '9') or (k == '-') or (k == '.')) and flagw == 1: 
                                    RefV += k
                                    flags = 0

                            if ('--' in Res) or Res == '.': 
                                Res = ""
                            if ('--' in RefV):
                                RefV = ""    
                                                            
                            flagw = 1
                            flags = 0

                            cell = sheet.cell(row=cnt, column=1, value=Res)
                            cell = sheet.cell(row=cnt, column=2, value=RefV)

                            Res = ""
                            RefV = ""
                            break
                    line = "" 

        
        
        workbook.save(file_path)
        workbook.close()
else:
    sys.exit(1)